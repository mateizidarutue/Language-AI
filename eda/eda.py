import sqlite3
from pathlib import Path

import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import webbrowser


FILE_PATH = Path("L&AI_data.xlsx")
DB_PATH = Path("lai.db")
CHUNK = 10_000  # reduce if memory is tight
OUTPUT_DIR = Path("eda_outputs")
SAMPLE_RATE = 0.02  # fraction of rows to sample for cross-sheet plots
SAMPLE_FOR_METRICS = 0.1  # fraction used for percentile calculations to keep memory reasonable
PAIR_SAMPLE_LIMIT = 5000  # max rows to sample for pairwise label agreements

# Sheet name -> label column name in the output DataFrame
SHEETS = {
    "sensing_intuitive": "sensing",
    "political_leaning": "political_leaning",
    "nationality": "nationality",
    "judging_perceiving": "judging_perceiving",
    "gender": "gender",
    "feeling_thinking": "feeling_thinking",
    "extrovert_introvert": "extrovert_introvert",
    "birth_year": "birth_year",
}


def load_sheet_to_sqlite(sheet_name: str, label_col: str, conn: sqlite3.Connection) -> None:
    """Stream a sheet into SQLite in chunks to keep memory low."""
    wb = load_workbook(FILE_PATH, read_only=True)
    ws = wb[sheet_name]

    # SQLite table schema: author_id TEXT, post TEXT, label TEXT
    conn.execute(f"DROP TABLE IF EXISTS [{sheet_name}]")
    conn.execute(
        f"""
        CREATE TABLE [{sheet_name}] (
            author_id TEXT,
            post TEXT,
            {label_col} TEXT
        )
        """
    )

    buffer = []
    total_rows = 0
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)  # skip header row
    for row in rows_iter:
        author_id, post, label = row[:3]
        buffer.append((author_id, post, label))
        if len(buffer) >= CHUNK:
            conn.executemany(
                f"INSERT INTO [{sheet_name}] (author_id, post, {label_col}) VALUES (?, ?, ?)",
                buffer,
            )
            total_rows += len(buffer)
            buffer = []

    if buffer:
        conn.executemany(
            f"INSERT INTO [{sheet_name}] (author_id, post, {label_col}) VALUES (?, ?, ?)",
            buffer,
        )
        total_rows += len(buffer)

    conn.commit()
    print(f"Loaded sheet '{sheet_name}' into table '{sheet_name}' with {total_rows} rows.")


def main() -> None:
    if not FILE_PATH.exists():
        raise FileNotFoundError(f"Could not find Excel file at {FILE_PATH}")

    OUTPUT_DIR.mkdir(exist_ok=True)

    with sqlite3.connect(DB_PATH) as conn:
        if _tables_exist(conn):
            print("Using existing SQLite tables; skipping reload.")
        else:
            for sheet, label_col in SHEETS.items():
                load_sheet_to_sqlite(sheet, label_col, conn)
            print(f"\nSQLite database written to {DB_PATH.resolve()}")
        metrics = run_basic_eda(conn)
        plot_table_sizes(metrics)
        plot_cross_sheet_lengths(conn)
        plot_entropy_balance(metrics)
        plot_length_vs_entropy(metrics)
        plot_metrics_correlation(metrics)
        plot_class_effective(metrics)
        plot_length_skew(metrics)
        write_summary_csv(metrics)
        author_coverage(conn)
        pairwise_label_agreement(conn)
    build_html_report()


def _tables_exist(conn: sqlite3.Connection) -> bool:
    """Return True if all expected tables already exist in the DB."""
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name IN ({})".format(
            ",".join(f"'{t}'" for t in SHEETS.keys())
        )
    )
    existing = {row[0] for row in cur.fetchall()}
    return existing == set(SHEETS.keys())


def run_basic_eda(conn: sqlite3.Connection) -> pd.DataFrame:
    """Print lightweight EDA summaries from the SQLite DB and return metrics."""
    print("\n--- Basic EDA ---")
    metrics_rows = []
    for sheet, label_col in SHEETS.items():
        total = conn.execute(f"SELECT COUNT(*) FROM [{sheet}]").fetchone()[0]
        missing = conn.execute(f"SELECT COUNT(*) FROM [{sheet}] WHERE {label_col} IS NULL").fetchone()[0]
        top_labels = conn.execute(
            f"""
            SELECT COALESCE({label_col}, 'MISSING') AS label, COUNT(*) AS n
            FROM [{sheet}]
            GROUP BY label
            ORDER BY n DESC
            LIMIT 5
            """
        ).fetchall()
        length_stats = conn.execute(
            f"""
            SELECT
                AVG(LENGTH(COALESCE(post,''))) AS avg_len,
                MIN(LENGTH(COALESCE(post,''))) AS min_len,
                MAX(LENGTH(COALESCE(post,''))) AS max_len
            FROM [{sheet}]
            """
        ).fetchone()
        print(f"\nSheet: {sheet}")
        print(f"  rows: {total:,} | missing {label_col}: {missing:,}")
        if total == 0:
            print("  table is empty; no stats available.")
            continue

        print("  top labels:")
        for label, n in top_labels:
            pct = n / total * 100 if total else 0
            print(f"    {label}: {n:,} ({pct:.2f}%)")
        majority_share = top_labels[0][1] / total if top_labels else 0
        label_entropy = _label_entropy(sheet, label_col, conn)
        print(f"  label majority share: {majority_share:.3f} | entropy: {label_entropy:.3f}")

        if length_stats and all(v is not None for v in length_stats):
            avg_len, min_len, max_len = length_stats
            print(f"  post length chars -> avg: {avg_len:.1f}, min: {min_len}, max: {max_len}")
        else:
            print("  post length stats: unavailable (empty or null data).")
            avg_len = min_len = max_len = None
        percentiles = _length_percentiles(sheet, conn)
        if percentiles:
            p_str = ", ".join(f"p{p}: {v}" for p, v in percentiles.items())
            print(f"  post length percentiles -> {p_str}")
        moments = _length_moments(sheet, conn)
        if moments:
            print(
                "  post length moments -> "
                f"std: {moments['std']:.1f}, skew: {moments['skew']:.3f}, kurtosis: {moments['kurt']:.3f}"
            )

        sample = conn.execute(
            f"""
            SELECT author_id, SUBSTR(post, 1, 120) AS snippet, {label_col}
            FROM [{sheet}]
            LIMIT 2
            """
        ).fetchall()
        print("  sample rows:")
        for row in sample:
            print(f"    {row}")

        plot_label_distribution(sheet, label_col, conn)
        plot_post_length(sheet, conn)

        metrics_rows.append(
            {
                "sheet": sheet,
                "label_col": label_col,
                "rows": total,
                "missing_labels": missing,
                "majority_share": majority_share,
                "label_entropy": label_entropy,
                "avg_len": avg_len,
                "min_len": min_len,
                "max_len": max_len,
                **{f"p{p}": percentiles.get(p) if percentiles else None for p in [25, 50, 75, 90, 95, 99]},
                "len_std": moments.get("std") if moments else None,
                "len_skew": moments.get("skew") if moments else None,
                "len_kurt": moments.get("kurt") if moments else None,
                "effective_classes": _effective_classes(sheet, label_col, conn),
            }
        )

    return pd.DataFrame(metrics_rows)


def plot_label_distribution(sheet: str, label_col: str, conn: sqlite3.Connection) -> None:
    """Save a bar chart of label distribution."""
    df = pd.read_sql_query(
        f"""
        SELECT COALESCE({label_col}, 'MISSING') AS label, COUNT(*) AS n
        FROM [{sheet}]
        GROUP BY label
        ORDER BY n DESC
        """,
        conn,
    )
    if df.empty:
        return
    plt.figure(figsize=(6, 4))
    plt.bar(df["label"], df["n"], color="#4C72B0")
    plt.title(f"{sheet} - {label_col} distribution")
    plt.ylabel("count")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    out_path = OUTPUT_DIR / f"{sheet}_labels.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_post_length(sheet: str, conn: sqlite3.Connection) -> None:
    """Save a histogram of post lengths."""
    lengths = pd.read_sql_query(
        f"SELECT LENGTH(COALESCE(post,'')) AS len FROM [{sheet}]",
        conn,
    )["len"]
    if lengths.empty:
        return
    capped = np.clip(lengths, None, np.percentile(lengths, 99))
    plt.figure(figsize=(6, 4))
    plt.hist(capped, bins=40, color="#55A868", edgecolor="white")
    plt.title(f"{sheet} - post length (chars, 99th pct capped)")
    plt.xlabel("chars")
    plt.ylabel("count")
    plt.tight_layout()
    out_path = OUTPUT_DIR / f"{sheet}_post_length.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_table_sizes(metrics: pd.DataFrame) -> None:
    """Bar chart of row counts per sheet."""
    if metrics.empty:
        return
    df = metrics[["sheet", "rows"]]
    plt.figure(figsize=(7, 4))
    plt.bar(df["sheet"], df["rows"], color="#8172B3")
    plt.title("Row counts per sheet")
    plt.ylabel("rows")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "all_table_sizes.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_cross_sheet_lengths(conn: sqlite3.Connection) -> None:
    """Compare post length distributions across sheets (sampled)."""
    lengths = {}
    for sheet in SHEETS.keys():
        sample = pd.read_sql_query(
            f"""
            SELECT LENGTH(COALESCE(post,'')) AS len
            FROM [{sheet}]
            WHERE ABS(RANDOM()) % 100 < {int(SAMPLE_RATE*100)}
            """,
            conn,
        )["len"]
        if not sample.empty:
            lengths[sheet] = np.clip(sample, None, np.percentile(sample, 99))
    if not lengths:
        return
    plt.figure(figsize=(8, 4))
    plt.boxplot(
        lengths.values(),
        labels=lengths.keys(),
        showfliers=False,
    )
    plt.ylabel("post length (chars, 99th pct capped)")
    plt.title(f"Post length comparison across sheets (sample ~{int(SAMPLE_RATE*100)}%)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "all_post_lengths.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def author_coverage(conn: sqlite3.Connection) -> None:
    """Report unique author counts per sheet and overlaps."""
    print("\n--- Author Coverage ---")
    author_sets = {}
    for sheet in SHEETS.keys():
        ids = pd.read_sql_query(f"SELECT DISTINCT author_id FROM [{sheet}]", conn)["author_id"]
        author_sets[sheet] = set(ids.dropna())
        print(f"  {sheet}: {len(author_sets[sheet]):,} unique authors")

    total_unique = len(set().union(*author_sets.values()))
    print(f"  total unique authors across all sheets: {total_unique:,}")

    # overlap heatmap (Jaccard similarity)
    sheets = list(SHEETS.keys())
    n = len(sheets)
    jaccard = np.zeros((n, n))
    for i, s1 in enumerate(sheets):
        for j, s2 in enumerate(sheets):
            if i == j:
                jaccard[i, j] = 1.0
            else:
                inter = len(author_sets[s1] & author_sets[s2])
                union = len(author_sets[s1] | author_sets[s2]) or 1
                jaccard[i, j] = inter / union
    plt.figure(figsize=(6, 5))
    im = plt.imshow(jaccard, cmap="viridis", vmin=0, vmax=1)
    plt.colorbar(im, fraction=0.046, pad=0.04, label="Jaccard similarity")
    plt.xticks(ticks=range(n), labels=sheets, rotation=45, ha="right")
    plt.yticks(ticks=range(n), labels=sheets)
    plt.title("Author overlap across sheets")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "author_overlap.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def _label_entropy(sheet: str, label_col: str, conn: sqlite3.Connection) -> float:
    """Shannon entropy of label distribution."""
    df = pd.read_sql_query(
        f"SELECT COALESCE({label_col}, 'MISSING') AS label, COUNT(*) AS n FROM [{sheet}] GROUP BY 1",
        conn,
    )
    total = df["n"].sum()
    if total == 0:
        return 0.0
    probs = df["n"] / total
    return float(-(probs * np.log2(probs)).sum())


def _length_percentiles(sheet: str, conn: sqlite3.Connection) -> dict:
    """Compute length percentiles on a sample for speed."""
    sample = pd.read_sql_query(
        f"""
        SELECT LENGTH(COALESCE(post,'')) AS len
        FROM [{sheet}]
        WHERE ABS(RANDOM()) % 100 < {int(SAMPLE_FOR_METRICS*100)}
        """,
        conn,
    )["len"]
    if sample.empty:
        return {}
    percentiles = [25, 50, 75, 90, 95, 99]
    values = np.percentile(sample, percentiles).astype(int)
    return {p: v for p, v in zip(percentiles, values)}


def _length_moments(sheet: str, conn: sqlite3.Connection) -> dict:
    """Compute std, skew, kurtosis on a sample of lengths."""
    sample = pd.read_sql_query(
        f"""
        SELECT LENGTH(COALESCE(post,'')) AS len
        FROM [{sheet}]
        WHERE ABS(RANDOM()) % 100 < {int(SAMPLE_FOR_METRICS*100)}
        """,
        conn,
    )["len"]
    if sample.empty:
        return {}
    return {
        "std": float(sample.std(ddof=0)),
        "skew": float(sample.skew()),
        "kurt": float(sample.kurtosis()),
    }


def _effective_classes(sheet: str, label_col: str, conn: sqlite3.Connection) -> float:
    """Effective number of classes: 1 / sum(p^2)."""
    df = pd.read_sql_query(
        f"SELECT COALESCE({label_col}, 'MISSING') AS label, COUNT(*) AS n FROM [{sheet}] GROUP BY 1",
        conn,
    )
    total = df["n"].sum()
    if total == 0:
        return 0.0
    probs = df["n"] / total
    return float(1 / (probs**2).sum())


def plot_entropy_balance(metrics: pd.DataFrame) -> None:
    """Plot label entropy and majority share across sheets."""
    if metrics.empty:
        return
    plt.figure(figsize=(8, 4))
    plt.bar(metrics["sheet"], metrics["label_entropy"], color="#CCB974", label="entropy")
    plt.plot(metrics["sheet"], metrics["majority_share"], color="#64B5CD", marker="o", label="majority share")
    plt.ylabel("entropy / share")
    plt.title("Label balance across sheets")
    plt.xticks(rotation=45, ha="right")
    plt.legend()
    plt.tight_layout()
    out_path = OUTPUT_DIR / "label_balance.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_length_vs_entropy(metrics: pd.DataFrame) -> None:
    """Scatter plot of avg post length vs label entropy."""
    if metrics.empty:
        return
    plt.figure(figsize=(6, 4))
    plt.scatter(metrics["label_entropy"], metrics["avg_len"], color="#E17C05")
    for _, row in metrics.iterrows():
        plt.text(row["label_entropy"], row["avg_len"], row["sheet"], fontsize=8, ha="right", va="bottom")
    plt.xlabel("label entropy")
    plt.ylabel("avg post length (chars)")
    plt.title("Avg post length vs label entropy")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "length_vs_entropy.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_metrics_correlation(metrics: pd.DataFrame) -> None:
    """Heatmap of correlations across numeric metrics."""
    if metrics.empty:
        return
    num_cols = ["rows", "missing_labels", "majority_share", "label_entropy", "avg_len", "len_std", "len_skew", "len_kurt"]
    df = metrics[num_cols].copy()
    corr = df.corr()
    plt.figure(figsize=(7, 5))
    im = plt.imshow(corr, cmap="coolwarm", vmin=-1, vmax=1)
    plt.colorbar(im, fraction=0.046, pad=0.04)
    plt.xticks(ticks=range(len(num_cols)), labels=num_cols, rotation=45, ha="right")
    plt.yticks(ticks=range(len(num_cols)), labels=num_cols)
    plt.title("Correlation of sheet-level metrics")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "metrics_correlation.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_class_effective(metrics: pd.DataFrame) -> None:
    """Bar chart of effective number of classes."""
    if metrics.empty or "effective_classes" not in metrics:
        return
    plt.figure(figsize=(8, 4))
    plt.bar(metrics["sheet"], metrics["effective_classes"], color="#86B6A5")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("effective classes (1/sum p^2)")
    plt.title("Effective number of classes per sheet")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "effective_classes.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def plot_length_skew(metrics: pd.DataFrame) -> None:
    """Plot skew/kurtosis of post lengths per sheet."""
    if metrics.empty:
        return
    plt.figure(figsize=(8, 4))
    plt.bar(metrics["sheet"], metrics["len_skew"], color="#E39C3C", label="skew")
    plt.plot(metrics["sheet"], metrics["len_kurt"], color="#7C7C7C", marker="o", label="kurtosis")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("skew / kurtosis")
    plt.title("Post length distribution shape")
    plt.legend()
    plt.tight_layout()
    out_path = OUTPUT_DIR / "length_shape.png"
    plt.savefig(out_path, dpi=150)
    plt.close()

def write_summary_csv(metrics: pd.DataFrame) -> None:
    """Write a CSV of per-sheet metrics for quick inspection."""
    if metrics.empty:
        return
    metrics.to_csv(OUTPUT_DIR / "summary.csv", index=False)


def author_coverage(conn: sqlite3.Connection) -> None:
    """Report unique author counts per sheet and overlaps."""
    print("\n--- Author Coverage ---")
    author_sets = {}
    for sheet in SHEETS.keys():
        ids = pd.read_sql_query(f"SELECT DISTINCT author_id FROM [{sheet}]", conn)["author_id"]
        author_sets[sheet] = set(ids.dropna())
        print(f"  {sheet}: {len(author_sets[sheet]):,} unique authors")

    total_unique = len(set().union(*author_sets.values()))
    print(f"  total unique authors across all sheets: {total_unique:,}")

    # overlap heatmap (Jaccard similarity)
    sheets = list(SHEETS.keys())
    n = len(sheets)
    jaccard = np.zeros((n, n))
    for i, s1 in enumerate(sheets):
        for j, s2 in enumerate(sheets):
            if i == j:
                jaccard[i, j] = 1.0
            else:
                inter = len(author_sets[s1] & author_sets[s2])
                union = len(author_sets[s1] | author_sets[s2]) or 1
                jaccard[i, j] = inter / union
    plt.figure(figsize=(6, 5))
    im = plt.imshow(jaccard, cmap="viridis", vmin=0, vmax=1)
    plt.colorbar(im, fraction=0.046, pad=0.04, label="Jaccard similarity")
    plt.xticks(ticks=range(n), labels=sheets, rotation=45, ha="right")
    plt.yticks(ticks=range(n), labels=sheets)
    plt.title("Author overlap across sheets")
    plt.tight_layout()
    out_path = OUTPUT_DIR / "author_overlap.png"
    plt.savefig(out_path, dpi=150)
    plt.close()


def pairwise_label_agreement(conn: sqlite3.Connection) -> None:
    """Compute pairwise label agreements for selected sheet pairs on shared authors."""
    pairs = [
        ("sensing_intuitive", "judging_perceiving"),
        ("sensing_intuitive", "feeling_thinking"),
        ("sensing_intuitive", "extrovert_introvert"),
        ("judging_perceiving", "feeling_thinking"),
        ("judging_perceiving", "extrovert_introvert"),
        ("feeling_thinking", "extrovert_introvert"),
    ]
    print("\n--- Pairwise label agreement (sampled) ---")
    for a, b in pairs:
        col_a = SHEETS[a]
        col_b = SHEETS[b]
        df = pd.read_sql_query(
            f"""
            SELECT a.author_id, a.{col_a} AS a_label, b.{col_b} AS b_label
            FROM [{a}] a
            INNER JOIN [{b}] b ON a.author_id = b.author_id
            WHERE a.author_id IS NOT NULL
            LIMIT {PAIR_SAMPLE_LIMIT}
            """,
            conn,
        )
        if df.empty:
            print(f"  {a} vs {b}: no overlap")
            continue
        contingency = pd.crosstab(df["a_label"], df["b_label"], normalize="all")
        agreement = (df["a_label"] == df["b_label"]).mean()
        print(f"  {a} vs {b}: overlap {len(df)} rows, agreement {agreement:.3f}")
        plt.figure(figsize=(5, 4))
        plt.imshow(contingency, cmap="Blues")
        plt.colorbar(label="proportion")
        plt.xticks(ticks=range(len(contingency.columns)), labels=contingency.columns, rotation=45, ha="right")
        plt.yticks(ticks=range(len(contingency.index)), labels=contingency.index)
        plt.title(f"{a} vs {b} (proportion)")
        plt.tight_layout()
        out_path = OUTPUT_DIR / f"{a}_vs_{b}_heatmap.png"
        plt.savefig(out_path, dpi=150)
        plt.close()


def build_html_report() -> None:
    """Build a simple HTML report referencing the generated plots and open it."""
    images = sorted(OUTPUT_DIR.glob("*.png"))
    summary_path = OUTPUT_DIR / "summary.csv"
    html_path = OUTPUT_DIR / "eda_report.html"

    img_tags = "\n".join(
        f'<div style="margin-bottom:20px;"><h3>{img.name}</h3><img src="{img.name}" style="max-width:100%;"></div>'
        for img in images
    )

    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>EDA Report</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 20px; }}
    img {{ border: 1px solid #ddd; }}
  </style>
</head>
<body>
  <h1>EDA Report</h1>
  <p>Summary CSV: <a href="{summary_path.name}">{summary_path.name}</a></p>
  {img_tags}
</body>
</html>
"""
    html_path.write_text(html, encoding="utf-8")
    print(f"\nHTML report written to {html_path.resolve()}")
    try:
        webbrowser.open(html_path.resolve().as_uri())
        print("Opened report in your default browser.")
    except Exception:
        print("Could not auto-open browser; open the HTML manually.")


if __name__ == "__main__":
    main()
